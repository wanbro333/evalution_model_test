from __future__ import annotations

import json
import re
import sys
import unittest
from html.parser import HTMLParser
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

MODEL_DIR = Path(__file__).resolve().parents[1]
if str(MODEL_DIR) not in sys.path:
    sys.path.insert(0, str(MODEL_DIR))

from config import BUDGET_PROBLEM_1, BUDGET_PROBLEM_2, OUTPUT_DIR
from credit_strategy import adjust_probability_by_odds, pava
from reporting import validate_workbook
from risk_model import grade_from_pd
from sensitivity import topsis_weight_perturbation


class _DashboardParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.ids = set()
        self.hrefs = []
        self.images = []

    def handle_starttag(self, tag, attrs):
        values = dict(attrs)
        if 'id' in values:
            self.ids.add(values['id'])
        if tag == 'a' and 'href' in values:
            self.hrefs.append(values['href'])
        if tag == 'img' and 'src' in values:
            self.images.append(values['src'])


class ModelUnitTests(unittest.TestCase):
    def test_pava_is_monotonic_and_preserves_length(self):
        fitted = pava([0.02, 0.04, 0.03, 0.08, 0.07])
        self.assertEqual(len(fitted), 5)
        self.assertTrue(np.all(np.diff(fitted) >= -1e-12))

    def test_pd_grade_boundaries_are_fixed(self):
        actual = grade_from_pd(np.array([0.0, 0.0999, 0.10, 0.20, 0.40, 0.99])).tolist()
        self.assertEqual(actual, ['A', 'A', 'B', 'C', 'D', 'D'])

    def test_odds_shock_is_bounded_and_directional(self):
        base = np.array([0.05, 0.20, 0.70])
        adverse = adjust_probability_by_odds(base, 1.30)
        positive = adjust_probability_by_odds(base, 0.90)
        self.assertTrue(np.all((adverse > base) & (adverse < 1)))
        self.assertTrue(np.all((positive < base) & (positive > 0)))


class GeneratedOutputTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        required = [
            OUTPUT_DIR / 'strategy_附件1.csv',
            OUTPUT_DIR / 'strategy_附件2.csv',
            OUTPUT_DIR / 'strategy_附件2_疫情调整.csv',
            OUTPUT_DIR / 'model_package.json',
            OUTPUT_DIR / '信贷决策建模结果.xlsx',
        ]
        missing = [str(path) for path in required if not path.exists()]
        if missing:
            raise unittest.SkipTest('请先运行 main.py：' + ', '.join(missing))
        cls.s1 = pd.read_csv(required[0])
        cls.s2 = pd.read_csv(required[1])
        cls.s3 = pd.read_csv(required[2])

    def _assert_allocation(self, frame, amount_col, budget):
        amount = frame[amount_col].to_numpy(float)
        self.assertAlmostEqual(float(amount.sum()), budget, places=4)
        positive = amount[amount > 1e-7]
        self.assertTrue(np.all(positive >= 10 - 1e-6))
        self.assertTrue(np.all(positive <= 100 + 1e-6))

    def test_allocation_constraints_and_d_exclusion(self):
        self._assert_allocation(self.s1, '贷款额度_万元', BUDGET_PROBLEM_1)
        self._assert_allocation(self.s2, '贷款额度_万元', BUDGET_PROBLEM_2)
        self._assert_allocation(self.s3, '调整后额度', BUDGET_PROBLEM_2)
        self.assertEqual(int((self.s1.loc[self.s1['信誉评级'] == 'D', '贷款额度_万元'] > 0).sum()), 0)
        self.assertEqual(int((self.s1.loc[self.s1['风险等级'] == 'D', '贷款额度_万元'] > 0).sum()), 0)
        self.assertEqual(int((self.s2.loc[self.s2['风险等级'] == 'D', '贷款额度_万元'] > 0).sum()), 0)
        self.assertEqual(int((self.s3.loc[self.s3['调整后等级'] == 'D', '调整后额度'] > 0).sum()), 0)

    def test_excel_contains_valid_formulas_and_correct_assumption_refs(self):
        path = OUTPUT_DIR / '信贷决策建模结果.xlsx'
        check = validate_workbook(path)
        self.assertGreater(check['formula_count'], 700)
        self.assertEqual(check['errors'], [])
        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb['问题1策略']
        heads = {cell.value: cell.column for cell in ws[1]}
        formula = ws.cell(2, heads['期望净收益']).value
        wb.close()
        self.assertIn("'模型说明与假设'!$B$7", formula)
        self.assertIn("'模型说明与假设'!$B$8", formula)
        self.assertNotIn("'模型说明与假设'!$B$6", formula)

    def test_sensitivity_is_recomputed_and_reproducible(self):
        features = pd.read_csv(OUTPUT_DIR / 'features_附件1.csv', index_col='企业代号')
        with (OUTPUT_DIR / 'model_package.json').open(encoding='utf-8') as handle:
            package = json.load(handle)
        first = topsis_weight_perturbation(features, package['topsis'], n_trials=5, seed=13)
        second = topsis_weight_perturbation(features, package['topsis'], n_trials=5, seed=13)
        pd.testing.assert_frame_equal(first, second)
        self.assertGreater(float(first['Spearman相关系数'].std()), 0)

    def test_dashboard_is_offline_and_uses_generated_assets(self):
        html = (MODEL_DIR / 'dashboard.html').read_text(encoding='utf-8')
        self.assertNotIn('http://', html)
        self.assertNotIn('https://', html)
        self.assertNotIn('—', html)
        self.assertNotIn('–', html)
        self.assertIn('output/charts/fig_default_rate.png', html)
        self.assertIn('output/charts/fig_data_scale.png', html)
        self.assertIn('output/charts/fig_profit_breakdown.png', html)
        self.assertIn('@media(max-width:640px)', html)
        self.assertIn('红冲示例', html)
        self.assertIn('为什么很多入选企业都是100万元', html)

        parser = _DashboardParser()
        parser.feed(html)
        required = {'model', 'data', 'validation', 'topsis', 'pricing', 'strategy', 'emergency', 'sensitivity', 'assumptions'}
        self.assertTrue(required.issubset(parser.ids))
        for href in parser.hrefs:
            if href.startswith('#'):
                self.assertIn(href[1:], parser.ids)
        for source in parser.images:
            self.assertTrue((MODEL_DIR / source).exists(), source)

    def test_paper_is_beginner_friendly_and_within_page_target(self):
        tex = (MODEL_DIR / 'paper' / 'paper.tex').read_text(encoding='utf-8')
        page_count = len(PdfReader(MODEL_DIR / 'paper' / 'paper.pdf').pages)
        self.assertGreaterEqual(page_count, 18)
        self.assertLessEqual(page_count, 22)
        self.assertGreaterEqual(tex.count('\\begin{intuitionbox}'), 8)
        self.assertGreaterEqual(tex.count('\\begin{decisionbox}'), 4)
        self.assertIn('主要符号说明', tex)
        self.assertIn('从原始发票到信贷策略的完整流程', tex)
        log_path = MODEL_DIR / 'paper' / 'paper.log'
        if log_path.exists():
            log = log_path.read_text(encoding='utf-8', errors='ignore')
            self.assertNotRegex(log, r'Overfull|Underfull|LaTeX Warning|Package .* Warning')


if __name__ == '__main__':
    unittest.main()
