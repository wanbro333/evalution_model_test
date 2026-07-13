"""统一生成指标JSON、专业Excel、静态Dashboard和LaTeX指标宏。"""
from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from html import escape
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import (
    BUDGET_PROBLEM_1,
    BUDGET_PROBLEM_2,
    DB_PATH,
    FUNDING_COST,
    GRADE_THRESHOLDS,
    LGD,
    LOAN_MAX,
    LOAN_MIN,
    MODEL_DIR,
    OUTPUT_DIR,
    PAPER_DIR,
    RANDOM_SEED,
    ensure_directories,
)


def _strategy_metrics(df, amount, rate, net):
    loan = df[df[amount] > 0]
    result = {
        'loan_count': int(len(loan)),
        'total_amount': float(loan[amount].sum()),
        'average_amount': float(loan[amount].mean()) if len(loan) else 0.0,
        'minimum_amount': float(loan[amount].min()) if len(loan) else 0.0,
        'maximum_amount': float(loan[amount].max()) if len(loan) else 0.0,
        'average_rate': float(loan[rate].mean()) if len(loan) else 0.0,
        'expected_net_profit': float(df[net].sum()),
    }
    for source, target in [
        ('预计实际放款额', 'expected_disbursement'),
        ('预计留存后利息', 'retained_interest'),
        ('预计违约损失', 'expected_default_loss'),
        ('预计资金成本', 'expected_funding_cost'),
    ]:
        result[target] = float(df[source].sum()) if source in df.columns else 0.0
    return result


def _database_summary():
    tables = {
        'attachment1_enterprises': 'enterprise_info_1',
        'attachment1_input_invoices': 'input_invoice_1',
        'attachment1_output_invoices': 'output_invoice_1',
        'attachment2_enterprises': 'enterprise_info_2',
        'attachment2_input_invoices': 'input_invoice_2',
        'attachment2_output_invoices': 'output_invoice_2',
        'rate_churn_rows': 'rate_churn',
    }
    connection = sqlite3.connect(DB_PATH)
    try:
        return {
            key: int(connection.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0])
            for key, table in tables.items()
        }
    finally:
        connection.close()


def build_report_metrics(scores1, scores2, strategy1, strategy2, strategy3, validation, sensitivity, churn_diagnostics, model_package):
    industry = strategy3.groupby('行业类别').agg(
        企业数=('企业代号', 'size'),
        调整前平均额度=('原贷款额度', 'mean'),
        调整后平均额度=('调整后额度', 'mean'),
        调整前平均利率=('原贷款利率', 'mean'),
        调整后平均利率=('调整后利率', 'mean'),
        调整前平均PD=('原违约概率', 'mean'),
        调整后平均PD=('调整后违约概率', 'mean'),
    ).reset_index()
    metrics = {
        'generated_at': datetime.now().astimezone().isoformat(timespec='seconds'),
        'model_version': '2.0-supervised-pd',
        'random_seed': RANDOM_SEED,
        'assumptions': {
            'problem1_budget_wan': BUDGET_PROBLEM_1,
            'problem2_budget_wan': BUDGET_PROBLEM_2,
            'lgd': LGD,
            'funding_cost': FUNDING_COST,
            'loan_min_wan': LOAN_MIN,
            'loan_max_wan': LOAN_MAX,
            'grade_thresholds': GRADE_THRESHOLDS,
        },
        'validation': validation,
        'data_summary': _database_summary(),
        'model': {
            'final_l2': model_package['lambda'],
            'feature_count': len(model_package['preprocessor']['features']),
            'attachment1_grade_distribution': {str(k): int(v) for k, v in scores1['风险等级'].value_counts().items()},
            'attachment2_grade_distribution': {str(k): int(v) for k, v in scores2['风险等级'].value_counts().items()},
        },
        'strategy1': _strategy_metrics(strategy1, '贷款额度_万元', '贷款年利率', '期望净收益'),
        'strategy2': _strategy_metrics(strategy2, '贷款额度_万元', '贷款年利率', '期望净收益'),
        'strategy3': _strategy_metrics(strategy3, '调整后额度', '调整后利率', '调整后期望净收益'),
        'original_d_loan_count': int(((strategy1.get('信誉评级') == 'D') & (strategy1['贷款额度_万元'] > 0)).sum()),
        'original_d_loan_amount': float(strategy1.loc[strategy1.get('信誉评级') == 'D', '贷款额度_万元'].sum()),
        'emergency_grade_changes': int((strategy3['原风险等级'] != strategy3['调整后等级']).sum()),
        'industry_summary': industry.to_dict('records'),
        'sensitivity': sensitivity,
        'churn_diagnostics': churn_diagnostics,
    }
    return metrics


def save_report_metrics(metrics):
    ensure_directories()
    path = OUTPUT_DIR / 'report_metrics.json'
    with open(path, 'w', encoding='utf-8') as handle:
        json.dump(metrics, handle, ensure_ascii=False, indent=2)
    return path


def write_paper_macros(metrics):
    grade = {row['风险等级']: row for row in metrics['validation']['grade_table']}
    commands = {
        'ModelAUC': f"{metrics['validation']['roc_auc']:.3f}",
        'ModelPRAUC': f"{metrics['validation']['pr_auc']:.3f}",
        'ModelBrier': f"{metrics['validation']['brier']:.3f}",
        'ModelLogLoss': f"{metrics['validation']['log_loss']:.3f}",
        'ModelLTwo': f"{metrics['model']['final_l2']:.2f}",
        'ModelFeatureCount': str(metrics['model']['feature_count']),
        'AttachmentOneEnterprises': str(metrics['data_summary']['attachment1_enterprises']),
        'AttachmentTwoEnterprises': str(metrics['data_summary']['attachment2_enterprises']),
        'AttachmentOneInvoices': str(metrics['data_summary']['attachment1_input_invoices'] + metrics['data_summary']['attachment1_output_invoices']),
        'AttachmentTwoInvoices': str(metrics['data_summary']['attachment2_input_invoices'] + metrics['data_summary']['attachment2_output_invoices']),
        'KnownDefaults': str(sum(row['违约数'] for row in metrics['validation']['grade_table'])),
        'KnownDefaultRate': f"{metrics['validation']['base_default_rate'] * 100:.1f}",
        'PoneLoanCount': str(metrics['strategy1']['loan_count']),
        'PtwoLoanCount': str(metrics['strategy2']['loan_count']),
        'PthreeLoanCount': str(metrics['strategy3']['loan_count']),
        'PoneAmount': f"{metrics['strategy1']['total_amount']:.0f}",
        'PtwoAmount': f"{metrics['strategy2']['total_amount']:.0f}",
        'PthreeAmount': f"{metrics['strategy3']['total_amount']:.0f}",
        'PoneAverageRate': f"{metrics['strategy1']['average_rate'] * 100:.2f}",
        'PtwoAverageRate': f"{metrics['strategy2']['average_rate'] * 100:.2f}",
        'PthreeAverageRate': f"{metrics['strategy3']['average_rate'] * 100:.2f}",
        'PoneNetProfit': f"{metrics['strategy1']['expected_net_profit']:.2f}",
        'PtwoNetProfit': f"{metrics['strategy2']['expected_net_profit']:.2f}",
        'PthreeNetProfit': f"{metrics['strategy3']['expected_net_profit']:.2f}",
        'EmergencyChanges': str(metrics['emergency_grade_changes']),
        'SensitivityMean': f"{metrics['sensitivity']['spearman_mean']:.4f}",
        'SensitivityMin': f"{metrics['sensitivity']['spearman_min']:.4f}",
        'SensitivityKendall': f"{metrics['sensitivity']['kendall_mean']:.4f}",
        'SensitivityIndicator': str(metrics['sensitivity']['most_influential_indicator']),
        'SensitivityRemoval': f"{metrics['sensitivity']['most_influential_removal_spearman']:.4f}",
        'OriginalDLoanCount': str(metrics['original_d_loan_count']),
    }
    for name in ['A', 'B', 'C', 'D']:
        row = grade.get(name, {'企业数': 0, '违约数': 0, '实际违约率': 0, '平均预测PD': 0})
        commands[f'Grade{name}Count'] = str(row['企业数'])
        commands[f'Grade{name}Defaults'] = str(row['违约数'])
        commands[f'Grade{name}Rate'] = f"{row['实际违约率'] * 100:.1f}"
        commands[f'Grade{name}PD'] = f"{row['平均预测PD'] * 100:.1f}"
    content = ['% 此文件由 main.py 自动生成，不要手工修改。']
    content.extend([f'\\newcommand{{\\{key}}}{{{value}}}' for key, value in commands.items()])
    path = PAPER_DIR / 'generated_metrics.tex'
    path.write_text('\n'.join(content) + '\n', encoding='utf-8')
    return path


def _excel_assumptions():
    return pd.DataFrame([
        ('模型版本', '2.0-supervised-pd', '监督逻辑回归PD，TOPSIS仅作解释'),
        ('生成时间', datetime.now().astimezone().isoformat(timespec='seconds'), '本次端到端运行时间'),
        ('随机种子', RANDOM_SEED, '交叉验证与灵敏度试验'),
        ('问题1预算(万元)', BUDGET_PROBLEM_1, '题目未给具体值，作为可配置情景'),
        ('问题2/3预算(万元)', BUDGET_PROBLEM_2, '题目给定1亿元'),
        ('违约损失率LGD', LGD, '可调整业务假设'),
        ('资金成本', FUNDING_COST, '可调整业务假设'),
        ('单户最低额度(万元)', LOAN_MIN, '题目约束'),
        ('单户最高额度(万元)', LOAN_MAX, '题目约束'),
        ('A级PD上限', GRADE_THRESHOLDS['A'][1], '不含上限'),
        ('B级PD上限', GRADE_THRESHOLDS['B'][1], '不含上限'),
        ('C级PD上限', GRADE_THRESHOLDS['C'][1], '不含上限'),
    ], columns=['参数', '数值', '说明'])


def _write_strategy_formulas(ws, adjusted=False):
    headers = {cell.value: cell.column for cell in ws[1]}
    if adjusted:
        mapping = {
            'amount': '调整后额度', 'rate': '调整后利率', 'pd': '调整后违约概率',
            'churn': '调整后预计流失率', 'net': '调整后期望净收益',
        }
    else:
        mapping = {
            'amount': '贷款额度_万元', 'rate': '贷款年利率', 'pd': '违约概率',
            'churn': '预计流失率', 'net': '期望净收益',
        }
    if not all(value in headers for value in mapping.values()):
        return
    amount = get_column_letter(headers[mapping['amount']])
    rate = get_column_letter(headers[mapping['rate']])
    pd_col = get_column_letter(headers[mapping['pd']])
    churn = get_column_letter(headers[mapping['churn']])
    net_col = headers[mapping['net']]
    for row in range(2, ws.max_row + 1):
        ws.cell(row, net_col).value = (
            f'={amount}{row}*(1-{churn}{row})*((1-{pd_col}{row})*{rate}{row}'
            f'-{pd_col}{row}*\'模型说明与假设\'!$B$7-\'模型说明与假设\'!$B$8)'
        )
        ws.cell(row, net_col).font = Font(name='Microsoft YaHei', color='000000')
        ws.cell(row, net_col).number_format = '0.00;[Red](0.00);-'


def export_excel(scores1, scores2, strategy1, strategy2, strategy3, coefficients, validation, sensitivity_trials, indicator_removal, output_path):
    ensure_directories()
    validation_summary = pd.DataFrame([
        ('ROC-AUC', validation['roc_auc']), ('PR-AUC', validation['pr_auc']),
        ('Brier分数', validation['brier']), ('Log Loss', validation['log_loss']),
        ('样本违约率', validation['base_default_rate']),
    ], columns=['指标', '数值'])
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _excel_assumptions().to_excel(writer, sheet_name='模型说明与假设', index=False)
        scores1.to_excel(writer, sheet_name='附件1风险评分', index=False)
        scores2.to_excel(writer, sheet_name='附件2风险评分', index=False)
        strategy1.to_excel(writer, sheet_name='问题1策略', index=False)
        strategy2.to_excel(writer, sheet_name='问题2策略', index=False)
        strategy3.to_excel(writer, sheet_name='问题3策略', index=False)
        validation_summary.to_excel(writer, sheet_name='模型验证', index=False)
        pd.DataFrame(validation['grade_table']).to_excel(writer, sheet_name='模型验证', index=False, startrow=8)
        coefficients.to_excel(writer, sheet_name='特征系数', index=False)
        sensitivity_trials.to_excel(writer, sheet_name='灵敏度试验', index=False)
        indicator_removal.to_excel(writer, sheet_name='指标移除', index=False)

    wb = load_workbook(output_path)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
    header_fill = PatternFill('solid', fgColor='17324D')
    assumption_fill = PatternFill('solid', fgColor='FFF2CC')
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Microsoft YaHei', size=10, color='000000')
                cell.alignment = Alignment(vertical='center')
        for column in ws.columns:
            values = [str(cell.value) if cell.value is not None else '' for cell in list(column)[:300]]
            width = min(max(max(map(len, values), default=8) * 1.25 + 2, 10), 32)
            ws.column_dimensions[get_column_letter(column[0].column)].width = width
        ws.auto_filter.ref = ws.dimensions
        if ws.max_row >= 2 and ws.max_column >= 1 and ws.title != '模型验证':
            safe_name = f"T{abs(hash(ws.title)) % 10**8}"
            table = Table(displayName=safe_name, ref=ws.dimensions)
            table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    assumptions = wb['模型说明与假设']
    for row in range(2, assumptions.max_row + 1):
        assumptions.cell(row, 2).fill = assumption_fill
        assumptions.cell(row, 2).font = Font(name='Microsoft YaHei', color='0000FF')
    for row in [7, 8, 11, 12, 13]:
        assumptions.cell(row, 2).number_format = '0.0%'

    _write_strategy_formulas(wb['问题1策略'])
    _write_strategy_formulas(wb['问题2策略'])
    _write_strategy_formulas(wb['问题3策略'], adjusted=True)

    summary = wb.create_sheet('结果汇总', 0)
    summary.append(['问题', '放贷企业数', '额度合计(万元)', '平均利率', '期望净收益(万元)'])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF')
    definitions = [
        ('问题1', '问题1策略', '贷款额度_万元', '贷款年利率', '期望净收益'),
        ('问题2', '问题2策略', '贷款额度_万元', '贷款年利率', '期望净收益'),
        ('问题3', '问题3策略', '调整后额度', '调整后利率', '调整后期望净收益'),
    ]
    for row_index, (label, sheet_name, amount_name, rate_name, net_name) in enumerate(definitions, start=2):
        ws = wb[sheet_name]
        heads = {cell.value: cell.column for cell in ws[1]}
        amount_letter = get_column_letter(heads[amount_name])
        rate_letter = get_column_letter(heads[rate_name])
        net_letter = get_column_letter(heads[net_name])
        end = ws.max_row
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, f'=COUNTIF(\'{sheet_name}\'!{amount_letter}2:{amount_letter}{end},">0")')
        summary.cell(row_index, 3, f'=SUM(\'{sheet_name}\'!{amount_letter}2:{amount_letter}{end})')
        summary.cell(row_index, 4, f'=AVERAGEIF(\'{sheet_name}\'!{amount_letter}2:{amount_letter}{end},">0",\'{sheet_name}\'!{rate_letter}2:{rate_letter}{end})')
        summary.cell(row_index, 5, f'=SUM(\'{sheet_name}\'!{net_letter}2:{net_letter}{end})')
    summary.freeze_panes = 'A2'
    summary.sheet_view.showGridLines = False
    for col, width in {'A': 14, 'B': 16, 'C': 20, 'D': 14, 'E': 22}.items():
        summary.column_dimensions[col].width = width
    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Microsoft YaHei', color='000000')
    for cell in summary['D'][1:]:
        cell.number_format = '0.0%'
    for col in ['C', 'E']:
        for cell in summary[col][1:]:
            cell.number_format = '0.00;[Red](0.00);-'
    wb.save(output_path)
    return output_path


def _grade_rows(metrics):
    rows = []
    for item in metrics['validation']['grade_table']:
        rows.append(
            f"<tr><td>{escape(item['风险等级'])}级</td><td>{item['企业数']}</td><td>{item['违约数']}</td>"
            f"<td>{item['实际违约率'] * 100:.1f}%</td><td>{item['平均预测PD'] * 100:.1f}%</td></tr>"
        )
    return ''.join(rows)


def _industry_rows(metrics):
    rows = []
    for item in metrics['industry_summary']:
        rows.append(
            f"<tr><td>{escape(str(item['行业类别']))}</td><td>{int(item['企业数'])}</td>"
            f"<td>{item['调整前平均额度']:.1f}</td><td>{item['调整后平均额度']:.1f}</td>"
            f"<td>{item['调整前平均PD'] * 100:.1f}%</td><td>{item['调整后平均PD'] * 100:.1f}%</td></tr>"
        )
    return ''.join(rows)


def _legacy_write_dashboard(metrics, path=MODEL_DIR / 'dashboard.html'):
    v = metrics['validation']
    s1, s2, s3 = metrics['strategy1'], metrics['strategy2'], metrics['strategy3']
    html = f'''<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>中小微企业信贷决策分析</title>
<style>
:root{{--bg:#f5f7f8;--surface:#ffffff;--ink:#14242d;--muted:#60727d;--line:#d9e2e7;--accent:#0f766e;--danger:#a83a32;--radius:12px}}
*{{box-sizing:border-box}}html{{scroll-behavior:smooth}}body{{margin:0;background:var(--bg);color:var(--ink);font-family:"Microsoft YaHei","Noto Sans CJK SC",Arial,sans-serif;line-height:1.55}}
nav{{position:sticky;top:0;z-index:10;height:64px;display:flex;align-items:center;justify-content:space-between;padding:0 max(24px,calc((100vw - 1180px)/2));background:#17324d;color:#f7fafb;border-bottom:1px solid #29475f}}
nav strong{{font-size:17px}}nav div{{display:flex;gap:22px}}nav a{{color:#c9d5dc;text-decoration:none;font-size:13px}}nav a:hover,nav a:focus{{color:#ffffff}}
main{{max-width:1180px;margin:auto;padding:42px 24px 80px}}header{{display:grid;grid-template-columns:1.4fr .6fr;gap:40px;align-items:end;padding:24px 0 34px;border-bottom:1px solid var(--line)}}
h1{{font-size:40px;line-height:1.2;margin:0 0 14px;letter-spacing:-1px}}header p{{margin:0;color:var(--muted);max-width:62ch}}.stamp{{text-align:right;color:var(--muted);font-size:12px;font-family:Consolas,monospace}}
.metrics{{display:grid;grid-template-columns:repeat(4,1fr);border-bottom:1px solid var(--line)}}.metric{{padding:24px 18px;border-right:1px solid var(--line)}}.metric:last-child{{border-right:0}}.metric b{{display:block;font:700 30px/1.1 Consolas,monospace;color:var(--accent)}}.metric span{{font-size:12px;color:var(--muted)}}
section{{padding:54px 0;border-bottom:1px solid var(--line)}}h2{{font-size:24px;margin:0 0 8px}}.lead{{margin:0 0 24px;color:var(--muted);max-width:70ch}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:22px}}figure{{margin:0;background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:16px}}figure img{{display:block;width:100%;height:auto}}figcaption{{font-size:12px;color:var(--muted);margin-top:10px}}
table{{width:100%;border-collapse:collapse;background:var(--surface);border:1px solid var(--line);font-size:13px}}th,td{{padding:11px 13px;text-align:right;border-bottom:1px solid var(--line)}}th:first-child,td:first-child{{text-align:left}}th{{background:#edf3f4;color:#38505d;font-weight:700}}tr:last-child td{{border-bottom:0}}.strategy{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px}}.strategy article{{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:20px}}.strategy h3{{margin:0 0 14px;font-size:16px}}.strategy dl{{display:grid;grid-template-columns:1fr auto;gap:8px;margin:0;font-size:13px}}dt{{color:var(--muted)}}dd{{margin:0;font-family:Consolas,monospace}}.ok{{color:var(--accent);font-weight:700}}footer{{padding:30px 24px;text-align:center;color:var(--muted);font-size:12px}}
@media(max-width:800px){{nav div{{display:none}}header,.grid{{grid-template-columns:1fr}}.stamp{{text-align:left}}.metrics{{grid-template-columns:1fr 1fr}}.metric:nth-child(2){{border-right:0}}.strategy{{grid-template-columns:1fr}}h1{{font-size:30px}}}}
</style></head><body>
<nav><strong>中小微企业信贷决策</strong><div><a href="#validation">模型验证</a><a href="#strategy">信贷策略</a><a href="#emergency">疫情调整</a><a href="#sensitivity">稳健性</a></div></nav>
<main><header><div><h1>监督违约概率驱动的信贷决策</h1><p>使用样本外交叉验证评估违约概率，以固定尺度TOPSIS解释经营质量，并在留存、违约损失和资金成本约束下优化额度与利率。</p></div><div class="stamp">MODEL 2.0<br>{escape(metrics['generated_at'])}</div></header>
<div class="metrics"><div class="metric"><b>{v['roc_auc']:.3f}</b><span>样本外 ROC-AUC</span></div><div class="metric"><b>{v['pr_auc']:.3f}</b><span>样本外 PR-AUC</span></div><div class="metric"><b>{v['brier']:.3f}</b><span>Brier 分数</span></div><div class="metric"><b>{metrics['original_d_loan_count']}</b><span>原始D级获贷企业数</span></div></div>
<section id="validation"><h2>违约概率验证</h2><p class="lead">四个PD档位的实际违约率来自重复分层样本外预测，不使用训练内概率。</p><div class="grid"><figure><img src="output/charts/fig_default_rate.png" alt="各风险等级样本外违约率"><figcaption>PD档位与实际违约率</figcaption></figure><table><thead><tr><th>等级</th><th>企业数</th><th>违约数</th><th>实际违约率</th><th>平均PD</th></tr></thead><tbody>{_grade_rows(metrics)}</tbody></table></div></section>
<section id="strategy"><h2>额度与利率优化</h2><p class="lead">预算约束作用于授信额度，单户额度为0或10-100万元。客户流失率、LGD和资金成本均进入目标函数。</p><div class="strategy">
<article><h3>问题1</h3><dl><dt>放贷企业</dt><dd>{s1['loan_count']}家</dd><dt>额度合计</dt><dd>{s1['total_amount']:.0f}万元</dd><dt>平均利率</dt><dd>{s1['average_rate']*100:.2f}%</dd><dt>期望净收益</dt><dd>{s1['expected_net_profit']:.2f}万元</dd></dl></article>
<article><h3>问题2</h3><dl><dt>放贷企业</dt><dd>{s2['loan_count']}家</dd><dt>额度合计</dt><dd>{s2['total_amount']:.0f}万元</dd><dt>平均利率</dt><dd>{s2['average_rate']*100:.2f}%</dd><dt>期望净收益</dt><dd>{s2['expected_net_profit']:.2f}万元</dd></dl></article>
<article><h3>问题3</h3><dl><dt>放贷企业</dt><dd>{s3['loan_count']}家</dd><dt>额度合计</dt><dd>{s3['total_amount']:.0f}万元</dd><dt>平均利率</dt><dd>{s3['average_rate']*100:.2f}%</dd><dt>期望净收益</dt><dd>{s3['expected_net_profit']:.2f}万元</dd></dl></article></div></section>
<section id="emergency"><h2>疫情赔率调整</h2><p class="lead">冲击系数乘到违约赔率，并重新运行评级、定价和额度优化。共有{metrics['emergency_grade_changes']}家企业等级发生变化。</p><div class="grid"><figure><img src="output/charts/fig_covid.png" alt="疫情调整前后行业平均贷款额度"><figcaption>行业平均额度的真实调整结果</figcaption></figure><table><thead><tr><th>行业</th><th>企业数</th><th>调整前额度</th><th>调整后额度</th><th>调整前PD</th><th>调整后PD</th></tr></thead><tbody>{_industry_rows(metrics)}</tbody></table></div></section>
<section id="sensitivity"><h2>真实灵敏度分析</h2><p class="lead">每次试验都重新计算TOPSIS得分和企业排名。没有使用预设均值或合成曲线。</p><div class="grid"><figure><img src="output/charts/fig_sensitivity.png" alt="真实权重扰动Spearman相关系数"><figcaption>Spearman均值 {metrics['sensitivity']['spearman_mean']:.4f}，最小值 {metrics['sensitivity']['spearman_min']:.4f}</figcaption></figure><figure><img src="output/charts/fig_weights.png" alt="逻辑回归标准化系数"><figcaption>监督PD模型主要特征方向</figcaption></figure></div></section>
</main><footer>结果由同一次端到端运行生成。统计源文件：output/report_metrics.json</footer></body></html>'''
    if '—' in html or '–' in html:
        raise AssertionError('Dashboard含有禁用的长破折号字符')
    path.write_text(html, encoding='utf-8')
    return path


def validate_workbook(path):
    wb = load_workbook(path, data_only=False, read_only=True)
    errors = []
    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith('='):
                    formula_count += 1
                if value in {'#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?'}:
                    errors.append(f'{ws.title}!{cell.coordinate}:{value}')
    wb.close()
    return {'formula_count': formula_count, 'errors': errors}
