"""提取所有文件内容"""
import os, glob, sys
sys.stdout.reconfigure(encoding='utf-8')

base = r'D:\shumo\evalation modei test'

# 1. 读取 docx
import zipfile, xml.etree.ElementTree as ET

def extract_docx_text(docx_path):
    """从 docx 提取文本"""
    with zipfile.ZipFile(docx_path) as z:
        xml_content = z.read('word/document.xml')
    tree = ET.fromstring(xml_content)
    ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
    paragraphs = []
    for p in tree.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
        texts = []
        for t in p.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
            if t.text:
                texts.append(t.text)
        if texts:
            paragraphs.append(''.join(texts))
    return '\n'.join(paragraphs)

# Find the docx file
for f in glob.glob(os.path.join(base, '*.docx')):
    if '~' not in f:
        docx_path = f
        break

print("="*80)
print("DOCX 文件内容:")
print("="*80)
text = extract_docx_text(docx_path)
with open(os.path.join(base, 'docx_content.txt'), 'w', encoding='utf-8') as f:
    f.write(text)
print(text[:5000])
print("...(完整内容已保存到 docx_content.txt)")

# 2. 读取 Excel 文件
import openpyxl

for pattern, desc in [('附件1*', '附件1'), ('附件2*', '附件2'), ('附件3*', '附件3')]:
    files = glob.glob(os.path.join(base, pattern))
    if files:
        f = [x for x in files if '~' not in x][0]
        print(f"\n{'='*80}")
        print(f"{desc}: {os.path.basename(f)}")
        print("="*80)
        wb = openpyxl.load_workbook(f, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\nSheet: {sn}, Rows: {ws.max_row}, Cols: {ws.max_column}")
            # Print headers
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
            print(f"列名: {headers}")
            # Print first 10 rows
            for r in range(2, min(12, ws.max_row+1)):
                row_data = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                print(f"  Row {r}: {row_data}")
        wb.close()
